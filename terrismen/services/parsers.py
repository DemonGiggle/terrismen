from __future__ import annotations

import mimetypes
import re
import subprocess
import uuid
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any

import fitz
import xlrd
from docx import Document as DocxDocument
from docx.document import Document as DocxDocumentType
from docx.oxml.table import CT_Tbl
from docx.oxml.text.paragraph import CT_P
from docx.table import Table
from docx.text.paragraph import Paragraph
from openpyxl import load_workbook

from terrismen.llm.base import ImageInput


class ParserError(RuntimeError):
    pass


@dataclass(slots=True)
class ParsedSource:
    locator: str
    content: str
    page_number: int | None = None
    metadata: dict[str, Any] = field(default_factory=dict)
    images: list[tuple[Path, ImageInput]] = field(default_factory=list)


def parse_document(file_path: Path, images_dir: Path) -> tuple[str, list[ParsedSource]]:
    suffix = file_path.suffix.lower()
    if suffix == ".pdf":
        return "pdf", _parse_pdf(file_path, images_dir)
    if suffix == ".docx":
        return "docx", _parse_docx(file_path, images_dir)
    if suffix == ".doc":
        return "doc", _parse_doc(file_path)
    if suffix == ".xlsx":
        return "xlsx", _parse_xlsx(file_path, images_dir)
    if suffix == ".xls":
        return "xls", _parse_xls(file_path)
    if suffix in {".txt", ".md", ".text"}:
        return "text", _parse_text(file_path)
    raise ParserError(f"Unsupported file type: {suffix or 'unknown'}")


def _write_image(images_dir: Path, stem: str, extension: str, blob: bytes) -> tuple[Path, ImageInput]:
    clean_extension = extension if extension.startswith(".") else f".{extension}"
    target = images_dir / f"{stem}-{uuid.uuid4().hex[:8]}{clean_extension}"
    target.write_bytes(blob)
    mime_type = mimetypes.guess_type(target.name)[0] or "application/octet-stream"
    return target, ImageInput(mime_type=mime_type, data=blob)


def _chunk_text(lines: list[str], *, prefix: str, page_start: int = 1, max_chars: int = 2800) -> list[ParsedSource]:
    chunks: list[ParsedSource] = []
    current: list[str] = []
    current_chars = 0
    line_start = 1
    line_number = 1
    chunk_number = page_start

    for raw_line in lines:
        line = raw_line.rstrip()
        current.append(line)
        current_chars += len(line) + 1
        if current_chars >= max_chars:
            locator = f"{prefix} {chunk_number} (lines {line_start}-{line_number})"
            chunks.append(ParsedSource(locator=locator, content="\n".join(current).strip(), page_number=chunk_number))
            chunk_number += 1
            current = []
            current_chars = 0
            line_start = line_number + 1
        line_number += 1

    if current:
        locator = f"{prefix} {chunk_number} (lines {line_start}-{line_number - 1})"
        chunks.append(ParsedSource(locator=locator, content="\n".join(current).strip(), page_number=chunk_number))

    return [chunk for chunk in chunks if chunk.content]


def _parse_pdf(file_path: Path, images_dir: Path) -> list[ParsedSource]:
    document = fitz.open(file_path)
    sources: list[ParsedSource] = []
    try:
        for page_index in range(document.page_count):
            page = document.load_page(page_index)
            text = page.get_text("text").strip()
            images: list[tuple[Path, ImageInput]] = []
            for image_index, image_ref in enumerate(page.get_images(full=True), start=1):
                xref = image_ref[0]
                image_info = document.extract_image(xref)
                extension = image_info.get("ext", "png")
                blob = image_info["image"]
                images.append(_write_image(images_dir, f"{file_path.stem}-p{page_index + 1}-img{image_index}", extension, blob))
            if text or images:
                sources.append(
                    ParsedSource(
                        locator=f"Page {page_index + 1}",
                        page_number=page_index + 1,
                        content=text,
                        metadata={"kind": "pdf"},
                        images=images,
                    )
                )
    finally:
        document.close()

    if not sources:
        raise ParserError("The PDF did not produce any readable text or images.")
    return sources


def _iter_docx_blocks(document: DocxDocumentType):
    for child in document.element.body.iterchildren():
        if isinstance(child, CT_P):
            yield Paragraph(child, document)
        elif isinstance(child, CT_Tbl):
            yield Table(child, document)


def _extract_docx_text(document: DocxDocumentType) -> list[str]:
    lines: list[str] = []
    for block in _iter_docx_blocks(document):
        if isinstance(block, Paragraph):
            text = block.text.strip()
            if not text:
                continue
            if block.style and block.style.name and block.style.name.startswith("Heading"):
                lines.append(f"# {text}")
            else:
                lines.append(text)
        else:
            for row in block.rows:
                row_values = [cell.text.strip() for cell in row.cells if cell.text.strip()]
                if row_values:
                    lines.append(" | ".join(row_values))
    return lines


def _extract_docx_images(document: DocxDocumentType, images_dir: Path, stem: str) -> list[tuple[Path, ImageInput]]:
    images: list[tuple[Path, ImageInput]] = []
    for rel in document.part.rels.values():
        if "image" not in rel.target_ref:
            continue
        blob = rel.target_part.blob
        suffix = Path(rel.target_ref).suffix or ".png"
        images.append(_write_image(images_dir, f"{stem}-docx-image", suffix, blob))
    return images


def _parse_docx(file_path: Path, images_dir: Path) -> list[ParsedSource]:
    document = DocxDocument(file_path)
    lines = _extract_docx_text(document)
    chunks = _chunk_text(lines, prefix="Chunk")
    images = _extract_docx_images(document, images_dir, file_path.stem)
    if not chunks and not images:
        raise ParserError("The DOCX did not produce any readable text or images.")
    if chunks and images:
        for index, image in enumerate(images):
            chunks[min(index, len(chunks) - 1)].images.append(image)
    elif images:
        chunks = [ParsedSource(locator="Chunk 1", content="", page_number=1, metadata={"kind": "docx"}, images=images)]
    for chunk in chunks:
        chunk.metadata.setdefault("kind", "docx")
    return chunks


def _parse_doc(file_path: Path) -> list[ParsedSource]:
    try:
        result = subprocess.run(
            ["antiword", str(file_path)],
            check=True,
            capture_output=True,
            text=True,
        )
    except FileNotFoundError as exc:
        raise ParserError("Legacy .doc parsing requires the `antiword` CLI to be installed.") from exc
    except subprocess.CalledProcessError as exc:
        raise ParserError(f"antiword failed to parse {file_path.name}: {exc.stderr.strip()}") from exc

    chunks = _chunk_text(result.stdout.splitlines(), prefix="Chunk")
    if not chunks:
        raise ParserError("The DOC file did not produce any readable text.")
    for chunk in chunks:
        chunk.metadata["kind"] = "doc"
    return chunks


def _extract_openpyxl_sheet_images(worksheet, images_dir: Path, stem: str) -> list[tuple[Path, ImageInput]]:
    extracted: list[tuple[Path, ImageInput]] = []
    for image_index, image in enumerate(getattr(worksheet, "_images", []), start=1):
        try:
            blob = image._data()
        except Exception:
            continue
        extracted.append(_write_image(images_dir, f"{stem}-{worksheet.title}-img{image_index}", ".png", blob))
    return extracted


def _rows_to_chunks(rows: list[str], *, label_prefix: str) -> list[ParsedSource]:
    chunks: list[ParsedSource] = []
    group_size = 40
    for start in range(0, len(rows), group_size):
        end = min(start + group_size, len(rows))
        locator = f"{label_prefix} rows {start + 1}-{end}"
        chunks.append(
            ParsedSource(
                locator=locator,
                page_number=(start // group_size) + 1,
                content="\n".join(rows[start:end]),
            )
        )
    return chunks


def _parse_xlsx(file_path: Path, images_dir: Path) -> list[ParsedSource]:
    workbook = load_workbook(file_path, data_only=True)
    sources: list[ParsedSource] = []
    for worksheet in workbook.worksheets:
        rows: list[str] = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
            cleaned = [str(cell).strip() for cell in row if cell not in (None, "")]
            if cleaned:
                rows.append(f"{row_index}: " + " | ".join(cleaned))
        sheet_chunks = _rows_to_chunks(rows, label_prefix=f"Sheet {worksheet.title}")
        images = _extract_openpyxl_sheet_images(worksheet, images_dir, file_path.stem)
        if sheet_chunks and images:
            sheet_chunks[0].images.extend(images)
        elif images:
            sheet_chunks = [ParsedSource(locator=f"Sheet {worksheet.title} rows 1-1", page_number=1, content="", images=images)]
        for chunk in sheet_chunks:
            chunk.metadata["kind"] = "xlsx"
            chunk.metadata["sheet"] = worksheet.title
        sources.extend(sheet_chunks)
    if not sources:
        raise ParserError("The Excel workbook did not produce any readable rows or images.")
    return sources


def _parse_xls(file_path: Path) -> list[ParsedSource]:
    workbook = xlrd.open_workbook(file_path)
    sources: list[ParsedSource] = []
    for sheet in workbook.sheets():
        rows: list[str] = []
        for row_index in range(sheet.nrows):
            cleaned = [str(sheet.cell_value(row_index, col)).strip() for col in range(sheet.ncols)]
            cleaned = [value for value in cleaned if value]
            if cleaned:
                rows.append(f"{row_index + 1}: " + " | ".join(cleaned))
        sheet_chunks = _rows_to_chunks(rows, label_prefix=f"Sheet {sheet.name}")
        for chunk in sheet_chunks:
            chunk.metadata["kind"] = "xls"
            chunk.metadata["sheet"] = sheet.name
        sources.extend(sheet_chunks)
    if not sources:
        raise ParserError("The XLS workbook did not produce any readable rows.")
    return sources


def _decode_text(blob: bytes) -> str:
    for encoding in ("utf-8-sig", "utf-8", "latin-1"):
        try:
            return blob.decode(encoding)
        except UnicodeDecodeError:
            continue
    raise ParserError("Could not decode the plaintext document.")


def _parse_text(file_path: Path) -> list[ParsedSource]:
    text = _decode_text(file_path.read_bytes())
    lines = [line for line in text.splitlines()]
    chunks = _chunk_text(lines, prefix="Chunk")
    if not chunks and text.strip():
        chunks = [ParsedSource(locator="Chunk 1", page_number=1, content=text.strip())]
    if not chunks:
        raise ParserError("The plaintext document is empty.")
    for chunk in chunks:
        chunk.metadata["kind"] = "text"
    return chunks
